import io
from typing import Optional

from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pydantic import BaseModel

import db

app = FastAPI(title="MiApp Backend")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

db.init_db()

OTP_STORE = {}


class AdminLoginReq(BaseModel):
    username: str
    password: str


class AdminCreateReq(BaseModel):
    username: str
    password: str


class ChangePasswordReq(BaseModel):
    username: str
    current_password: str
    new_password: str


class WorkerCreateReq(BaseModel):
    full_name: str
    area: str
    email: Optional[str] = None
    face_key: Optional[str] = None


class WorkerUpdateReq(BaseModel):
    full_name: str
    area: str
    email: Optional[str] = None
    face_key: Optional[str] = None


class LogCreateReq(BaseModel):
    worker_id: int
    action: str
    log_type: str = "worker"
    method: str = "Rostro"


class SendOtpReq(BaseModel):
    worker_id: int
    action: str


class VerifyOtpReq(BaseModel):
    worker_id: int
    action: str
    code: str


class VisitorCreateReq(BaseModel):
    full_name: str
    company: Optional[str] = None
    phone: Optional[str] = None
    reason: Optional[str] = None
    photo_uri: Optional[str] = None


def fake_send_email(to_email: str, code: str):
    print(f"[OTP DEMO] Enviando código {code} a {to_email}")


@app.get("/health")
def health():
    return {"ok": True}


@app.post("/admin/login")
def admin_login(body: AdminLoginReq):
    admin = db.admin_login(body.username, body.password)
    if not admin:
        raise HTTPException(status_code=401, detail="Credenciales incorrectas")
    return {"ok": True, "admin": {"username": admin["username"]}}


@app.get("/admins")
def get_admins():
    try:
        return db.list_admins()
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/admins")
def post_admin(body: AdminCreateReq):
    try:
        return db.create_admin(body.username, body.password)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/admin/change-password")
def post_change_password(body: ChangePasswordReq):
    try:
        db.change_admin_password(
            username=body.username,
            current_password=body.current_password,
            new_password=body.new_password,
        )
        return {"ok": True}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/workers")
def get_workers():
    try:
        return db.list_workers()
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/workers")
def post_worker(body: WorkerCreateReq):
    try:
        return db.create_worker(
            full_name=body.full_name,
            area=body.area,
            email=body.email,
            face_key=body.face_key,
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/workers/{worker_id}")
def put_worker(worker_id: int, body: WorkerUpdateReq):
    try:
        updated = db.update_worker(
            worker_id=worker_id,
            full_name=body.full_name,
            area=body.area,
            email=body.email,
            face_key=body.face_key,
        )

        if not updated:
            raise HTTPException(status_code=404, detail="Trabajador no encontrado")

        return updated
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/workers/{worker_id}")
def delete_worker(worker_id: int):
    try:
        deleted = db.delete_worker(worker_id)
        if not deleted:
            raise HTTPException(status_code=404, detail="Trabajador no encontrado")
        return {"ok": True}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/logs")
def get_logs():
    try:
        return db.list_logs()
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/logs")
def post_log(body: LogCreateReq):
    try:
        return db.create_log(
            worker_id=body.worker_id,
            action=body.action,
            log_type=body.log_type,
            method=body.method,
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/auth/send-otp")
def send_otp(body: SendOtpReq):
    try:
        workers = db.list_workers()
        worker = next((w for w in workers if int(w["id"]) == int(body.worker_id)), None)

        if not worker:
            raise HTTPException(status_code=404, detail="Trabajador no encontrado")

        if not worker.get("email"):
            raise HTTPException(status_code=400, detail="El trabajador no tiene correo registrado")

        code = "123456"
        OTP_STORE[str(body.worker_id)] = {
            "code": code,
            "action": body.action,
        }

        fake_send_email(worker["email"], code)

        return {"ok": True, "message": "Código enviado", "demo_code": code}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/auth/verify-otp")
def verify_otp(body: VerifyOtpReq):
    try:
        saved = OTP_STORE.get(str(body.worker_id))

        if not saved:
            raise HTTPException(status_code=400, detail="No hay código generado para este trabajador")

        if saved["action"] != body.action:
            raise HTTPException(status_code=400, detail="La acción no coincide con el código generado")

        if str(saved["code"]) != str(body.code).strip():
            raise HTTPException(status_code=400, detail="Código incorrecto")

        result = db.create_log(
            worker_id=body.worker_id,
            action=body.action,
            log_type="worker",
            method="Correo",
        )

        del OTP_STORE[str(body.worker_id)]
        return {"ok": True, "log": result}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/visitors")
def get_visitors():
    try:
        return db.list_visitors()
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/visitors")
def post_visitor(body: VisitorCreateReq):
    try:
        return db.create_visitor(
            full_name=body.full_name,
            company=body.company,
            phone=body.phone,
            reason=body.reason,
            photo_uri=body.photo_uri,
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/export/asistencia.xlsx")
def export_asistencia():
    try:
        rows = db.list_logs()

        wb = Workbook()
        ws = wb.active
        ws.title = "Asistencia"

        headers = [
            "Nombre",
            "No. Empleado",
            "Departamento",
            "Tipo",
            "Método",
            "Fecha y Hora",
        ]
        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True, color="000000")
            cell.fill = PatternFill("solid", fgColor="D9D9D9")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for r in rows:
            ws.append([
                r.get("full_name", ""),
                r.get("employee_no", ""),
                r.get("area", ""),
                r.get("action", ""),
                r.get("method", "Rostro"),
                r.get("created_at", ""),
            ])

        widths = {
            "A": 22,
            "B": 16,
            "C": 20,
            "D": 14,
            "E": 14,
            "F": 24,
        }

        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)

        return StreamingResponse(
            bio,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="asistencia.xlsx"'},
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))